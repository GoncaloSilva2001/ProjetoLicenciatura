import csv
import emoji
import requests
from bs4 import BeautifulSoup
from openpyxl import load_workbook
from openpyxl.workbook import Workbook
from nltk.sentiment import SentimentIntensityAnalyzer

# Função para buscar emojis do site
def fetch_emojis(url):
    response = requests.get(url)
    if response.status_code == 200:
        soup = BeautifulSoup(response.content, 'html.parser')
        emojis = soup.find_all('span', class_='emoji emoji-large')
        emoji_list = [emoji.text.strip() for emoji in emojis]
        return emoji_list
    else:
        print(f"Failed to fetch emojis. Status code: {response.status_code}")
        return []

# Função para obter os pontos de código Unicode de um caractere
def get_unicode_code_points(string):
    return ' '.join([f'U+{ord(char):X}' for char in string])

# Função para obter a descrição de um emoji
def get_emoji_description(emoji_char):
    shortcode = emoji.demojize(emoji_char)
    description = shortcode.strip(':').replace('_', ' ')
    return description

# Função para calcular o score de análise de sentimentos usando NLTK
def calcular_score(descricao):
    analyzer = SentimentIntensityAnalyzer()
    scores = analyzer.polarity_scores(descricao)
    return scores['compound']

# Função para ler slang do arquivo e adicionar ao xlsx
def add_slangs_to_xlsx(filename, worksheet_name, slang_file):
    wb = load_workbook(filename)
    ws = wb[worksheet_name]

    with open(slang_file, 'r', encoding='utf-8') as file:
        reader = csv.reader(file, delimiter='=')
        for row in reader:
            if len(row) == 2:
                slang = row[1].strip()  # Remove leading and trailing whitespace
                if slang.endswith(','):  # Check if slang ends with a comma
                    slang = slang[:-1]  # Remove the trailing comma
                score = calcular_score(slang)  # Calcular o score de análise de sentimentos
                ws.append([row[0].strip(), '', slang, score])

    wb.save(filename)

if __name__ == "__main__":
    url = "https://getemoji.com/"
    emojis = fetch_emojis(url)
    data_for_xlsx = [['Design', 'Unicode', 'Description', 'Score']]

    for emoji_char in emojis:
        unicode_points = get_unicode_code_points(emoji_char)
        description = get_emoji_description(emoji_char)
        score = calcular_score(description)  # Calcular o score de análise de sentimentos
        data_for_xlsx.append([emoji_char, unicode_points, description, score])

    # Create a new Excel workbook
    wb = Workbook()
    ws = wb.active
    ws.title = "Emojis"

    # Write data to worksheet
    for row in data_for_xlsx:
        ws.append(row)

    # Save workbook to XLSX file
    wb.save("emojis_with_sentiment_score.xlsx")

    # Add slang terms to the xlsx
    add_slangs_to_xlsx("emojis_with_sentiment_score.xlsx", "Emojis", "slang.txt")

    print("Os ficheiros CSV e XLSX foram preenchidos com sucesso!")
